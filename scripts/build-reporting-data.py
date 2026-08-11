#!/usr/bin/env python3
"""Create the immutable 2026 reporting snapshot used by the dashboard.

Usage:
  build-reporting-data.py <deals.xlsx> <leads.xlsx> <output.json>

The dashboard intentionally reads only this generated file. It does not call
NetHunt or connect to the database for reporting metrics.
"""

from __future__ import annotations

import json
import sys
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


YEAR = "2026"


def text(value: Any) -> str | None:
    if value is None:
        return None
    value = str(value).strip()
    return value or None


def number(value: Any) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value or "").replace(" ", "").replace(",", "."))
    except ValueError:
        return 0.0


def iso_datetime(value: Any) -> str | None:
    if not isinstance(value, datetime):
        return None
    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone.utc)
    return value.astimezone(timezone.utc).isoformat().replace("+00:00", "Z")


def iso_day(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return None


def read_records(path: Path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "Records" not in workbook.sheetnames:
        raise ValueError(f"{path.name}: missing Records sheet")
    sheet = workbook["Records"]
    rows = sheet.iter_rows(values_only=True)
    headers = [text(cell) or "" for cell in next(rows)]
    return headers, rows


def make_deals(path: Path) -> list[dict[str, Any]]:
    headers, rows = read_records(path)
    index = {header: position for position, header in enumerate(headers)}
    required = ["<Record ID>", "<Created At>", "Name", "Воронка", "Бюджет"]
    missing = [name for name in required if name not in index]
    if missing:
        raise ValueError(f"{path.name}: missing deal columns {missing}")

    def at(row: tuple[Any, ...], name: str) -> Any:
        position = index.get(name)
        return row[position] if position is not None and position < len(row) else None

    deals: list[dict[str, Any]] = []
    for row in rows:
        record_id = text(at(row, "<Record ID>"))
        created_at = iso_datetime(at(row, "<Created At>"))
        if not record_id or not created_at or not created_at.startswith(YEAR):
            continue
        won_at = iso_day(at(row, "Дата - Виграні"))
        lost_at = iso_day(at(row, "Дата - Програні"))
        deals.append(
            {
                "id": record_id,
                "name": text(at(row, "Name")),
                "company": text(at(row, "Компанія")),
                "stage": text(at(row, "Воронка")) or "Запит",
                "clientType": text(at(row, "Тип клієнта")),
                "budget": number(at(row, "Бюджет")),
                "rawSource": text(at(row, "(Джерело) Where do they know us from?")),
                "channel": text(at(row, "(Канал) How did they contact us?")),
                "lossReason": text(at(row, "Причина програшу угоди")),
                "requestedAt": iso_day(at(row, "Дата - Запит")),
                "wonAt": won_at,
                "lostAt": lost_at,
                "createdAt": created_at,
                "createdMonth": created_at[:7],
                "closeMonth": (won_at or lost_at or "")[:7] or None,
            }
        )
    return deals


def make_leads(path: Path) -> list[dict[str, Any]]:
    headers, rows = read_records(path)
    index = {header: position for position, header in enumerate(headers)}
    required = ["<Record ID>", "<Created At>", "Воронка", "Дата створення"]
    missing = [name for name in required if name not in index]
    if missing:
        raise ValueError(f"{path.name}: missing lead columns {missing}")

    def at(row: tuple[Any, ...], name: str) -> Any:
        position = index.get(name)
        return row[position] if position is not None and position < len(row) else None

    leads: list[dict[str, Any]] = []
    for row in rows:
        record_id = text(at(row, "<Record ID>"))
        created_at = iso_datetime(at(row, "<Created At>"))
        lead_day = iso_day(at(row, "Дата створення")) or (created_at or "")[:10]
        if not record_id or not created_at or not lead_day or not lead_day.startswith(YEAR):
            continue
        updated_at = iso_datetime(at(row, "<Updated At>"))
        stage = text(at(row, "Воронка"))
        handling_hours: float | None = None
        if created_at and updated_at and stage and stage != "Новий":
            start = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
            end = datetime.fromisoformat(updated_at.replace("Z", "+00:00"))
            hours = (end - start).total_seconds() / 3600
            if 0 <= hours < 24 * 90:
                handling_hours = round(hours, 6)
        contact = text(at(row, "Ім'я")) or text(at(row, "Name"))
        if not contact:
            contact = text(at(row, "Телефон")) or text(at(row, "Email адреса"))
        leads.append(
            {
                "id": record_id,
                "contactName": contact,
                "stage": stage,
                "qualification": text(at(row, "Квал.")),
                "channel": text(at(row, "Канал (як зв'язались)")) or "Не вказано",
                "rawSource": text(at(row, "Джерело")),
                "utmSource": text(at(row, "utm_source")),
                "utmMedium": text(at(row, "utm_medium")),
                "utmCampaign": text(at(row, "utm_campaign")),
                "leadDate": lead_day,
                "createdAt": created_at,
                "updatedAt": updated_at,
                "month": lead_day[:7],
                "handlingHours": handling_hours,
            }
        )
    return leads


def main() -> None:
    if len(sys.argv) != 4:
        raise SystemExit("Usage: build-reporting-data.py <deals.xlsx> <leads.xlsx> <output.json>")
    deals_path, leads_path, output_path = map(Path, sys.argv[1:])
    deals = make_deals(deals_path)
    leads = make_leads(leads_path)
    payload = {
        "generatedAt": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "period": {"from": "2026-01", "to": "2026-12"},
        "sources": {"deals": deals_path.name, "leads": leads_path.name},
        "deals": deals,
        "leads": leads,
    }
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(f"Generated {output_path}: {len(deals)} deals, {len(leads)} leads")


if __name__ == "__main__":
    main()
